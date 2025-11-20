"""
    封装方法
    1.封装任意读取Excel文件，其中的表单可以指定读取
    2.数据写入
"""
import openpyxl


class HandleExel:

    def __init__(self, filename,sheet):
        """
        初始化Excel
        :param filename: Excel文件名
        :param sheet: 表单名
        """
        self.filename = filename
        self.sheet = sheet

    def get_exl_data(self):
        """
        # 读取数据
        :return:返回读取数据
        """
        # 读取excel
        wb = openpyxl.load_workbook(self.filename)
        # 读取表单
        sh = wb[self.sheet]
        # 获取表单所有行数据
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        case = []
        for item in res[1:]:
            ls = [j.value for j in item]
            dic = dict(zip(title, ls))
            case.append(dic)
        return case

    def write_exl_data(self, row, column, value):
        """
        # 写入数据
        :param row:写入的行
        :param column:写入的列
        :param value:参数
        :return:
        """
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheet]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)
